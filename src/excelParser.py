import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import OrderedDict
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import griddata
from scipy.io import savemat
from scipy.optimize import differential_evolution
import sys, os,json
# from pykrige.ok3d import OrdinaryKriging3D
# from pykrige.uk3d import UniversalKriging3D
import matlab.engine

def main(path, paramList):
    if sys.argv:
        if sys.argv[1] == 'run':
            dataFrame = csv2Dataframe(path)
            df2MsExcel(dataFrame,'trafficPatternType',paramList)
            df1 = pd.read_excel('results.xlsx', engine='openpyxl')
            data = []
            data_element = []
            for param in paramList:
                dfDict = df2Dict('results.xlsx',param)
                json2mat(dfDict, param)
                for key in dfDict:
                    for element in dfDict[key]:
                        data_element.append(int(element))
                    data.append(data_element)
            with open('data.json', 'w', encoding='utf-8') as f:
                json.dump(dfDict, f, ensure_ascii=False, indent=4)
            with open('data.json', 'w', encoding='utf-8') as f:
                json.dump(dfDict, f, ensure_ascii=False, indent=4)
            #dseMatModule() 
        elif sys.argv[1] == 'clean':
            os.system('rm -rf ./result.xlsx')
        else:
            print(sys.argv[1])
            print('use run/clean ')

    
def csv2Dataframe(path):
    '''
    reads results.csv from PANACA tool rearranges data and adds average latency
    column and returns the dataframe object
    return type : pd.DataFrame()
    '''
    df = pd.read_csv( path, encoding='utf-8', header = 0, sep=";", dtype={
                     "bufferSize": int, "packetLength": int, "Minimum latency":str,
                     "Maximum latency": str, "Average Latency":str, "trafficPattern": str,
                     "packetInjectionRate":float})
    df1 = df[['packetLength','bufferSize', 'Minimum latency', 'Maximum latency', 'Average latency','trafficPattern','packetInjectionRate']]

    df1['avgLatency'] = timeUnitNormalize(df1)
    df1['timePerFlit'] = df1['avgLatency']/(df1['packetLength'])
    df1['trafficPatternType'] = df1['trafficPattern'].str.extract(r'(\w*$)')[0].astype(str)
    df2 = df1[['Average latency','bufferSize', 'packetLength','packetInjectionRate','trafficPatternType','Minimum latency','Maximum latency','avgLatency']]
    return df2
def timeUnitNormalize(df1):
    '''
    Takes in Column of Dataframe and normalizes the unit and sends 
    back the column
    return : df.col
    '''
    df1['avgLatUnit'] =df1['Average latency'].str.extract('(\w*$)').astype(str)
    df1['avgLatWoUnit'] = df1['Average latency'].str.extract('(\d+)').astype(float)
    df1['avgLatency'] = df1['Average latency']
    df1['avgLatency'][df1.avgLatUnit == 's'] = df1['avgLatWoUnit']*1000000000.0
    df1['avgLatency'][df1.avgLatUnit == 'us'] = df1['avgLatWoUnit']*1000.0
    df1['avgLatency'][df1.avgLatUnit == 'ms'] = df1['avgLatWoUnit']*1000000.0
    df1['avgLatency'][df1.avgLatUnit == 'ns'] = df1['avgLatWoUnit']*1.0

    return df1['avgLatency']
def df2MsExcel(dataFrame, headerToConv,headerList2Sheet):
    '''
    takes dataframe and sorts by column cell value provided and
    saves them into multiple sheet excel Workbook
    '''
    data2Extract = []
    wb = Workbook()
    for headerName in dataFrame.columns:
        if headerName != headerToConv:
            data2Extract.append(headerName)
    for header2Sheet in headerList2Sheet:
        df2Wrt = dataFrame[dataFrame[headerToConv] == header2Sheet]
        df2Wrt = df2Wrt.sort_values(by=['packetLength','bufferSize'])
        df2Wrt.reset_index()      
        ws = wb.create_sheet(header2Sheet)
        dataFrame2Conv = searchColDuplicate(df2Wrt,['packetLength','bufferSize'])
        
        for r_idx, row in enumerate(dataframe_to_rows(dataFrame2Conv, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save('results.xlsx')
    wb = load_workbook('results.xlsx')
    ws = wb['Sheet']
    
    for r_idx, row in enumerate(dataframe_to_rows(dataFrame, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    # print(dataFrame)
    wb.save('results.xlsx')
      
      
def searchColDuplicate(dataFrame, paramList):
    '''
    Takes in pandas dataframe and groups and mean by the parameter
    list provided and returns the updated dataframe
    return : pandas Dataframe
    '''
    df1 = dataFrame[['packetLength','bufferSize','avgLatency']]
    df2 = pd.DataFrame()
    df2['packetLength'] = pd.to_numeric(df1['packetLength'])
    df2['bufferSize'] = pd.to_numeric(df1['bufferSize'])
    #df2['packetInjectionRate'] = pd.to_numeric(df1['packetInjectionRate'])
    df2['avgLatency'] = pd.to_numeric(df1['avgLatency'])
    mean_df = df2.groupby(paramList, as_index=False).mean()
    mean_df['timePerFlit'] = mean_df['avgLatency']/(mean_df['packetLength'])
    
    return mean_df
 
def df2Dict(path,sheet):
    '''
    read_data will take the path of the excel/csv/text data
    file and return an Ordered Dictionary with the data wrapped 
    inside it
    '''
    dfToDict = OrderedDict()
    dataFrame = pd.read_excel(path, engine='openpyxl', sheet_name = sheet,header=0, dtype={
      "avgLatency": float,
      "packetLength": float,
      "bufferSize": float,
      "timePerFlit": float

    })
    bufferSizeList   = dataFrame['bufferSize'].to_list()
    packetLengthList = dataFrame['packetLength'].to_list()
    timePerFlitList      = dataFrame['timePerFlit'].to_list()

    dfToDict['bufferSize'] = bufferSizeList
    dfToDict['packetLength'] = packetLengthList
    dfToDict['timePerFlit'] = timePerFlitList
    return dfToDict

def dictToJson(dictToConvert):
  '''
  converts Dictionary to Json and writes a JSON data file
  '''
  json_object = json.dumps(dictToConvert, indent = 4)
  with open("data.json", "w") as outfile:
    outfile.write(json_object)

def meshGrid2Interpolate(jsonData):
    xq,yq = np.mgrid[0:1:600, 0:1:128]
    
    bufferSize = np.array(jsonData['bufferSize']).T
    packetLength = np.array(jsonData['packetLength']).T
    timePerFlit = np.array(jsonData['timePerFlit']).T

    xy = np.zeros((2,np.size(bufferSize)))
    xy[0] = bufferSize
    xy[1] = packetLength
    xy = xy.T
    x_array = np.arange(0,601)
    y_array = np.arange(0,128)

    grid_x, grid_y = np.mgrid[0:601:1, 0:128:1]
    grid_z = griddata(xy, timePerFlit, (grid_x, grid_y), method='nearest')

    min_timeperflit = np.argmin(np.argmin(grid_z))
    # plot
    # fig = plt.figure()
    # ax = fig.add_subplot(111)
    # plt.contourf(bufferSize,timePerFlit,grid_z,np.arange(0,601,1))
    # plt.plot(bufferSize,timePerFlit,'k.')
    # plt.xlabel('xi',fontsize=16)
    # plt.ylabel('yi',fontsize=16)
    # plt.savefig('interpolated.png',dpi=100)
    # plt.close(fig)
    
def json2mat(jsonData, var_name):
    '''
    Takes in a python dictionary as input and converts and saves
    the struct matlab data type
    return : None
    '''
    dse_array = OrderedDict()
    bufferSize = np.array(jsonData['bufferSize']).T
    packetLength = np.array(jsonData['packetLength']).T
    timePerFlit = np.array(jsonData['timePerFlit']).T   
    dse_array['bufferSize'] = bufferSize
    dse_array['packetLength'] = packetLength
    dse_array['timePerFlit'] = timePerFlit
    basePath = './db/mat/'
    save_as = basePath + var_name +'.mat'
    dbMatExist = os.path.exists('./db/mat')
    if not dbMatExist:
        os.makedirs(basePath)
    savemat(save_as, dse_array)

def meshgrid2(*arrs):
  arrs = tuple(reversed(arrs))  #edit
  lens = list(map(len, arrs))
  dim = len(arrs)

  sz = 1
  for s in lens:
    sz*=s

  ans = []    
  for i, arr in enumerate(arrs):
    slc = [1]*dim
    slc[i] = lens[i]
    arr2 = np.asarray(arr).reshape(slc)
    for j, sz in enumerate(lens):
      if j!=i:
        arr2 = arr2.repeat(sz, axis=j) 
    ans.append(arr2)
def de(fobj, bounds, mut=0.8, crossp=0.7, popsize=20, its=1000):
    dimensions = len(bounds)
    pop = np.random.rand(popsize, dimensions)
    min_b, max_b = np.asarray(bounds).T
    diff = np.fabs(min_b - max_b)
    pop_denorm = min_b + pop * diff
    fitness = np.asarray([fobj(ind) for ind in pop_denorm])
    best_idx = np.argmin(fitness)
    best = pop_denorm[best_idx]
    for i in range(its):
        for j in range(popsize):
            idxs = [idx for idx in range(popsize) if idx != j]
            a, b, c = pop[np.random.choice(idxs, 3, replace = False)]
            mutant = np.clip(a + mut * (b - c), 0, 1)
            cross_points = np.random.rand(dimensions) < crossp
            if not np.any(cross_points):
                cross_points[np.random.randint(0, dimensions)] = True
            trial = np.where(cross_points, mutant, pop[j])
            trial_denorm = min_b + trial * diff
            f = fobj(trial_denorm)
            if f < fitness[j]:
                fitness[j] = f
                pop[j] = trial
                if f < fitness[best_idx]:
                    best_idx = j
                    best = trial_denorm
        yield best, fitness[best_idx]
def dseMatModule():
    eng = matlab.engine.start_matlab()
    dseVal = eng.dseMod(nargout=0)
    #eng.quit()
    
# def kriging(data2):
#     data = np.array([[0.1, 0.1, 0.3, 0.9],
#     [0.2, 0.1, 0.4, 0.8],
#     [0.1, 0.3, 0.1, 0.9],
#     [0.5, 0.4, 0.4, 0.5],
#     [0.3, 0.3, 0.2, 0.7]])
#     gridx = np.arange(0.0, 0.6, 0.01)
#     gridy = np.arange(0.0, 0.6, 0.01)
#     gridz = np.arange(0.0, 0.6, 0.1)

#     # Create the 3D ordinary kriging object and solves for the three-dimension kriged
#     # volume and variance. Refer to OrdinaryKriging3D.__doc__ for more information.
#     # ok3d = OrdinaryKriging3D(data[:, 0], data[:, 1], data[:, 2], data[:, 3],variogram_model='linear')
#     # k3d, ss3d = ok3d.execute('grid', gridx, gridy, gridz)
#     # Create the 3D universal kriging object and solves for the three-dimension kriged
#     # volume and variance. Refer to UniversalKriging3D.__doc__ for more information.
#     uk3d = UniversalKriging3D(data[:, 0], data[:, 1], data[:, 2], data[:, 3], variogram_model='linear', drift_terms=['regional_linear'])
#     k3d, ss3d = uk3d.execute('grid', gridx, gridy, gridz)
#     # To use the generic 'specified' drift term, the user must provide the drift values
#     # at each data point and at every grid point. The following example is equivalent to
#     # using a linear drift in all three spatial dimensions. Refer to
#     # UniversalKriging3D.__doc__ for more information.
#     zg, yg, xg = np.meshgrid(gridz, gridy, gridx, indexing='ij')
#     uk3d = UniversalKriging3D(data[:, 0], data[:, 1], data[:, 2], data[:, 3],variogram_model='linear', drift_terms=['specified'],
#     specified_drift=[data[:, 0], data[:, 1]])
#     k3d, ss3d = uk3d.execute('grid', gridx, gridy, gridz, specified_drift_arrays=[xg, yg,zg])
#     # To use the generic 'functional' drift term, the user must provide a callable
#     # function that takes only the spatial dimensions as arguments. The following example
#     # is equivalent to using a linear drift only in the x-direction. Refer to
#     # UniversalKriging3D.__doc__ for more information.
#     func = lambda x, y, z: x
#     uk3d = UniversalKriging3D(data[:, 0], data[:, 1], data[:, 2], data[:, 3],variogram_model='linear', drift_terms=['functional'],functional_drift=[func])
#     k3d, ss3d = uk3d.execute('grid', gridx, gridy, gridz)
#     # Note that the use of the 'specified' and 'functional' generic drift capabilities is
#     # essentially identical in the two-dimensional universal kriging class (except for a
#     # difference in the number of spatial coordinates for the passed drift functions).
#     # See UniversalKriging.__doc__ for more information.
def ackley(x):
    arg1 = -0.2 * np.sqrt(0.5 * (x[0] ** 2 + x[1] ** 2))
    arg2 = 0.5 * (np.cos(2. * np.pi * x[0]) + np.cos(2. * np.pi * x[1]))
    return -20. * np.exp(arg1) - np.exp(arg2) + 20. + np.e


    
if __name__ == '__main__':
    main('results.csv',['Bitcomplement', 'Bitrotate', 'Bitrevers', 'Transpose1', 'Transpose2', 'Bitshuffle'])
    dfDict = df2Dict('results.xlsx','Bitrotate')
    #json2mat(dfDict, 'Bitcomplement')
    dseMatModule()
    # bounds = [(-5, 5), (-5, 5)]
    # result = differential_evolution(ackley, bounds)
    # result.x, result.fun
    # print(result.x, result.fun)
  