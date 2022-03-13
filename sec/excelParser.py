import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import OrderedDict
import numpy as np
import matplotlib.pyplot as plt
from scipy.interpolate import griddata
from scipy.io import savemat

def main():
    dataFrame = csv2Dataframe('results.csv')
    df2MsExcel(dataFrame,'trafficPatternType',['Bitcomplement', 'Bitrevers', 'Bitrotate','Bitshuffle','Transpose1', 'Transpose2'])
    dfDict = df2Dict('results.xlsx','Bitcomplement')
    meshGrid2Interpolate(dfDict)


def csv2Dataframe(path):
    '''
    reads results.csv from PANACA tool rearranges data and adds average latency
    column and returns the dataframe object
    return type : pd.DataFrame()
    '''
    df = pd.read_csv( path, encoding='utf-8', header = 0, sep=";", dtype={
                     "bufferSize": int, "packetLength": int, "Minimum latency":str,
                     "Maximum latency": str, "Average Latency":str, "trafficPattern": str,
                     "packetInjectionRate":float})
    df1 = df[['bufferSize', 'packetLength', 'Minimum latency', 'Maximum latency', 'Average latency','trafficPattern','packetInjectionRate']]
    df1['avgLatency'] = df1['Average latency'].str.extract('(\d+)').astype(float)
    df1['timePerFlit'] = df1['avgLatency']/(df1['packetLength'])
    df1['trafficPatternType'] = df1['trafficPattern'].str.extract(r'(\w*$)')[0].astype(str)
    df2 = df1[['bufferSize', 'packetLength','packetInjectionRate','trafficPatternType','Minimum latency','Maximum latency','avgLatency']]
    return df2

def df2MsExcel(dataFrame, headerToConv,headerList2Sheet):
    '''
    takes dataframe and sorts by column cell value provided and
    saves them into multiple sheet excel Workbook
    '''
    data2Extract = []
    wb = Workbook()
    for headerName in dataFrame.columns:
        if headerName != headerToConv:
            data2Extract.append(headerName)
    for header2Sheet in headerList2Sheet:
        df2Wrt = dataFrame[dataFrame[headerToConv] == header2Sheet]
        df2Wrt = df2Wrt.sort_values(by=['bufferSize','packetLength'])
        df2Wrt.reset_index()      
        ws = wb.create_sheet(header2Sheet)
        dataFrame2Conv = searchColDuplicate(df2Wrt,['packetLength','bufferSize'])
        for r_idx, row in enumerate(dataframe_to_rows(dataFrame2Conv, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        wb.save('results.xlsx')
    wb = load_workbook('results.xlsx')
    ws = wb['Sheet']
    for r_idx, row in enumerate(dataframe_to_rows(dataFrame, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save('results.xlsx')
      
      
def searchColDuplicate(dataFrame, paramList):
    '''
    Takes in pandas dataframe and groups and mean by the parameter
    list provided and returns the updated dataframe
    return : pandas Dataframe
    '''
    df1 = dataFrame[['bufferSize','packetLength','avgLatency']]
    mean_df1 = df1.groupby(paramList, as_index=False).mean()
    mean_df1['timePerFlit'] = mean_df1['avgLatency']/(mean_df1['packetLength'])
    return mean_df1
 
def df2Dict(path,sheet):
    '''
    read_data will take the path of the excel/csv/text data
    file and return an Ordered Dictionary with the data wrapped 
    inside it
    '''
    dfToDict = OrderedDict()
    dataFrame = pd.read_excel(path, sheet_name = sheet,header=0, dtype={
      "avgLatency": float,
      "packetLength": float,
      "bufferSize": float,
      "timePerFlit": float

    })
    bufferSizeList   = dataFrame['bufferSize'].to_list()
    packetLengthList = dataFrame['packetLength'].to_list()
    timePerFlitList      = dataFrame['timePerFlit'].to_list()

    dfToDict['bufferSize'] = bufferSizeList
    dfToDict['packetLength'] = packetLengthList
    dfToDict['timePerFlit'] = timePerFlitList
    return dfToDict

def dictToJson(dictToConvert):
  '''
  converts Dictionary to Json and writes a JSON data file
  '''
  json_object = json.dumps(dictToConvert, indent = 4)
  with open("data.json", "w") as outfile:
    outfile.write(json_object)

def meshGrid2Interpolate(jsonData):
    xq,yq = np.mgrid[0:1:600, 0:1:128]
    
    bufferSize = np.array(jsonData['bufferSize']).T
    packetLength = np.array(jsonData['packetLength']).T
    timePerFlit = np.array(jsonData['timePerFlit']).T

    xy = np.zeros((2,np.size(bufferSize)))
    xy[0] = bufferSize
    xy[1] = packetLength
    xy = xy.T
    x_array = np.arange(0,601)
    y_array = np.arange(0,128)

    #grid_x, grid_y = np.meshgrid(x_array, y_array)
    #grid_x, grid_y = np.mgrid[0.0:0.09:601*1j, 0.0:0.03:128*1j]
    grid_x, grid_y = np.mgrid[0:601:1, 0:128:1]
    # print(grid_x)
    # print(grid_y)
    # print('XY')
    # print(xy)
    # print('Timeperflit')
    # print(timePerFlit)
    # print('Grid Z')
    grid_z = griddata(xy, timePerFlit, (grid_x, grid_y), method='nearest')

    dse_array = OrderedDict()
    dse_array['bufferSize'] = bufferSize
    dse_array['packetLength'] = packetLength
    dse_array['timePerFlit'] = timePerFlit
    dse_array['vq'] = grid_z
    savemat("dse_array.mat", dse_array)

    print(grid_z)
    min_timeperflit = np.argmin(np.argmin(grid_z))
    print(min_timeperflit)
    # plot
    # fig = plt.figure()
    # ax = fig.add_subplot(111)
    # plt.contourf(bufferSize,timePerFlit,grid_z,np.arange(0,601,1))
    # plt.plot(bufferSize,timePerFlit,'k.')
    # plt.xlabel('xi',fontsize=16)
    # plt.ylabel('yi',fontsize=16)
    # plt.savefig('interpolated.png',dpi=100)
    # plt.close(fig)

def meshgrid2(*arrs):
  arrs = tuple(reversed(arrs))  #edit
  lens = list(map(len, arrs))
  dim = len(arrs)

  sz = 1
  for s in lens:
    sz*=s

  ans = []    
  for i, arr in enumerate(arrs):
    slc = [1]*dim
    slc[i] = lens[i]
    arr2 = np.asarray(arr).reshape(slc)
    for j, sz in enumerate(lens):
      if j!=i:
        arr2 = arr2.repeat(sz, axis=j) 
    ans.append(arr2)


if __name__ == '__main__':
    main()
